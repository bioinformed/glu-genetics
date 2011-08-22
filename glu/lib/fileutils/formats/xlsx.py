# -*- coding: utf-8 -*-

__abstract__  = 'fileutils XLSX file support'
__copyright__ = 'Copyright (c) 2007-2009, BioInformed LLC and the U.S. Department of Health & Human Services. Funded by NCI under Contract N01-CO-12400.'
__license__   = 'See GLU license for terms by running: glu license'
__revision__  = '$Id$'


from   glu.lib.fileutils.auto   import namefile
from   glu.lib.fileutils.parser import parse_augmented_filename, tryint1, get_arg


def table_reader_xlsx(filename,strdata=True,extra_args=None,**kwargs):
  '''
  Load rows from a Microsoft Excel (XLSX) file using the openpyxl module

  Supports XLSX files produced by Excel versions 2007 and newer
  '''
  try:
    from openpyxl.reader.excel import load_workbook
  except ImportError:
    raise ValueError('Missing openpyxl module to read Microsoft Excel XLSX files')

  if extra_args is None:
    args = kwargs
  else:
    args = extra_args
    args.update(kwargs)

  name  = parse_augmented_filename(filename,args)
  sheet = get_arg(args, ['sheet'],0)
  hyin  = get_arg(args, ['hyphen'])

  if extra_args is None and args:
    raise ValueError('Unexpected filename arguments: %s' % ','.join(sorted(args)))

  if hyin is not None and name=='-':
    raise IOError('Cannot read Excel file from stdin')

  book = load_workbook(name, use_iterators=True)

  try:
    sheet = book.get_sheet_by_name(sheet)
  except KeyError:
    pass

  if sheet is None:
    try:
      #names = book.get_sheet_names()
      #name  = names[tryint1(sheet or 0)]
      #sheet = book.get_sheet_by_name(name)
      sheet = book.worksheets[tryint1(sheet or 0)]
    except (ValueError,IndexError):
      raise ValueError('Cannot open Excel sheet %s:%s' % (namefile(name),sheet))

  def _table_reader_xlsx(sheet):
    def convert(x):
      if isinstance(x,float):
        y = int(x)
        return str(y) if x==y else str(x)
      elif isinstance(x,unicode):
        return str(x)
      else:
        return x or ''

    for row in sheet.iter_rows():
      yield [ convert(cell.internal_value) for cell in row ]

  return _table_reader_xlsx(sheet)


class XLSXWriter(object):
  '''
  Write a Microsoft Excel (XLSX) file using the openpyxl module

  Supports XLSX files produced by Excel versions 2007 and newer
  '''
  def __init__(self, filename, extra_args=None, **kwargs):
    '''
    @param filename: file name or file object
    @type  filename: str or file object
    @param    sheet: sheet name
    @type     sheet: str
    '''
    try:
      from openpyxl.workbook import Workbook
    except ImportError:
      raise ValueError('Missing openpyxl module to write Microsoft Excel XLSX files')

    if extra_args is None:
      args = kwargs
    else:
      args = extra_args
      args.update(kwargs)

    name  = parse_augmented_filename(filename,args)

    sheet = get_arg(args, ['sheet'])
    hyout = get_arg(args, ['hyphen'])

    if extra_args is None and args:
      raise ValueError('Unexpected filename arguments: %s' % ','.join(sorted(args)))

    self.filename = name if name!='-' or hyout is None else hyout
    self.book     = Workbook(optimized_write=True)
    self.sheet    = self.book.create_sheet()
    if sheet:
      self.sheet.title = sheet

  def writerows(self, rows):
    '''
    Write a sequence of rows as rows in an XLSX file

    @param rows: rows to write to XLSX
    @type  rows: sequence of sequences of str
    '''
    if self.book is None:
      raise IOError('Writer object already closed')

    append  = self.sheet.append

    for row in rows:
      append(row)

  def writerow(self, row):
    '''
    Write a sequence of strings to a row in an XLSX file

    @param row: row to write to XLSX
    @type  row: sequences of str
    '''
    if self.book is None:
      raise IOError('Writer object already closed')

    self.sheet.append(row)

  def close(self):
    '''
    Close the writer

    A closed writer cannot be used for further I/O operations and will
    result in an error if called more than once.
    '''
    if self.book is None:
      raise IOError('Writer object already closed')

    self.book,book = None,self.book

    book.save(self.filename)

  def __enter__(self):
    '''
    Context enter function
    '''
    return self

  def __exit__(self, *exc_info):
    '''
    Context exit function that closes the writer upon exit
    '''
    self.close()

  def __del__(self):
    '''
    Finish saving output when the writer is destroyed, if not already saved.
    '''
    if getattr(self,'book',None) is not None:
      self.close()


def _test():
  import doctest
  return doctest.testmod()


if __name__ == '__main__':
  _test()
